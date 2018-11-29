#          STEPS TO RUN THIS REPORT - INCLUDING NEXT STEPS
###########################################################################################################
#       To successfully run this report, carefully review comments number 1, 2, 3, 4                        #
#       This script will spit out a single XLSX file, with each clinician having their own sheet.           # 
#                                                                                                           #
#       You can very easily split the Excel notebook into individual files using a VBA script found here:   #
#       https://www.extendoffice.com/documents/excel/628-excel-split-workbook.html                          #
############################################################################################################

#   Import required packages
import pandas as pd
from bs4 import BeautifulSoup
import numpy as np
from openpyxl.chart import StockChart, Reference, Series
from openpyxl.chart.axis import ChartLines
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font
from openpyxl import Workbook
from openpyxl.chart.data_source import NumDataSource, NumData, NumVal, NumRef
from openpyxl.styles import Alignment
from openpyxl.chart.error_bar import ErrorBars

from openpyxl.chart.series_factory import SeriesFactory

#   Increase the number of columns displayed for development purposes
pd.set_option('display.max_columns', 10)

def clean_HTML(text):
    #   Clean the HTML code using BS4
    soup = BeautifulSoup(text, "html.parser")
    get_soup = soup.get_text()
    return get_soup 

def list_plus_minus(ser):
    #   Find Min Max by Question
    ser_list = ser.tolist()
    results_list = sorted(ser_list)
    return results_list[0], results_list[-1]


def list2errorbars(plus, minus, errDir='y',errValType='cust'):
    #   Returns ErrorBar from lists of error values

    #   Convert to list of NumVal
    numval_plus = [NumVal(i, None, v=x) for i, x in enumerate(plus)]
    numval_minus = [NumVal(i, None, v=x) for i, x in enumerate(minus)]

    #   Convert to NumData
    nd_plus = NumData(pt=numval_plus)
    nd_minus = NumData(pt=numval_minus)

    #   Convert to NumDataSource
    nds_plus = NumDataSource(numLit=nd_plus)
    nds_minus = NumDataSource(numLit=nd_minus)

    return ErrorBars(plus=nds_plus, minus=nds_minus, errDir=errDir, errValType=errValType)



######################################################################################
#       1                                                                            #
#       THIS IS WHERE YOU HAVE TO CHANGE THE LINK TO THE NEWLY DOWNLOADED REPORT     #
######################################################################################
data = pd.read_csv(r"C:\Users\ARidding\Documents\Clevelands\Cleveland_Data - 1of4 - 2018-2019.csv", header=0)
######################################################################################

#   Remove all negative values
non_na_data = data[data["Answer"] >= 0 ]

#   Store the minimum and max values
plus_group = non_na_data.groupby(["Question_ID"])

#   Store the min/max of the answer column, by question_id
plus = non_na_data.groupby(["Question_ID"])["Answer"].max().tolist()
minus = non_na_data.groupby(["Question_ID"])["Answer"].min().tolist()


#Create a Pivot table by mean score
clean_df = data[["Evaluation_Name","Response_ID","Question","Question_ID","Answer","Comments","Clinician_AD_ID"]]

s = clean_df['Answer']

#  Working only with numeric data
mask = s.isin(['0', '1', '2', '3', '4', '5'])


#   Filter out the questions, convert to numeric, clean data
clean_df = clean_df[mask]
s = pd.to_numeric(clean_df.Answer)
clean_df["Numeric"] = s

#   Drop the Answer Column
clean_df = clean_df.drop("Answer", 1)

#   Create pivot tables to aggregate data
clean_pivot = clean_df.pivot_table(index=["Question_ID"],columns=["Clinician_AD_ID"], values=["Numeric"], aggfunc='mean')
glo_pivot = clean_df.pivot_table(index=["Question_ID"], values=["Numeric"], aggfunc='mean')

#   Add Global Average to the clean pivot table.
clean_pivot["Global Mean"] = glo_pivot

#   Create a copy not to mess with the original
clean_data = data.copy()
clean_data = clean_data[['Evaluation_Name',"Question_ID",'Question',"Clinician_AD_ID","Comments"]]

#   Remove the HTML Code from the Comments field
clean_question = clean_data["Question"].apply(clean_HTML)

#   Keep all columns except the last column
clean_data = clean_data.iloc[:-1,:]

#   Create a new column "Question" with the clean questions we just stripped
clean_data["Clean_Question"] = clean_question

#   Create unique list of questions   
d_quest = clean_data[["Question_ID","Clean_Question"]]
d_quest = d_quest.drop_duplicates()

#   Set the index to Question_ID
d_quest = d_quest.set_index('Question_ID')

#   Unique question list
quest_uni = d_quest.Clean_Question.unique()

#   Create mask
mask = np.ones(len(quest_uni),dtype=bool)

# #######################################################################################################
# #       2                                                                                             #
# #       The two "Comment" based questions (non-numeric) will need to be removed from the chart        #
# #       Starting from 0, COUNT THE NUMBER OF THESE TWO QUESTIONS AND PLACE THEM IN THE BRACKETS BELOW #
# #######################################################################################################
mask[[0,1]] = False                                                                                #
# #######################################################################################################

#   Remove questions which are not relevant (comment questions)
question_result = quest_uni[mask]
question_res = pd.Series(question_result)

# ###########################################################################
# #    3                                                                    #
# #    CONFIRM THAT THE ORDER OF THE QUESTIONS ARE CORRECT IN THE GRAPH     #
# ###########################################################################
l_order = [0, 13, 4, 10, 11, 15, 12, 14, 3 ,8, 2, 7, 6, 1, 9, 5]                        #
# ###########################################################################

#   Set the question order
question_res.index = l_order
question_res = question_res.sort_index()

#   Add Min/Max to clean_pivot
clean_pivot["Min"] = minus
clean_pivot["Max"] = plus

#   Sort df by Order Colum
clean_pivot["Order"] = l_order
clean_pivot = clean_pivot.sort_values(by=["Order"])#

#   Retrieve the sorted Min/Max
sorted_minus = []
sorted_plus = []


#   Average across the set
av_across_set = clean_pivot["Global Mean"].mean()*20

#########################   Short Answer Grouping ##########################################
#   Scrub HTML from the data export
data["Question"] = data["Question"].apply(clean_HTML)

#   Filter out the short answer questions
short_answer_data = data[data["Answer"]==-2]

#   Apply groupby, first by clinician and then by Question ID. 
short_answer_sort = short_answer_data.sort_values(["Clinician_AD_ID","Question_ID"])


for _val in range(0, len(clean_pivot)):
    glob_av_per_question = clean_pivot.iloc[_val,-4] * 20
    question_minimum = (clean_pivot.iloc[_val,:-4].min() * 20)
    question_minimum = glob_av_per_question - question_minimum
    question_max = (clean_pivot.iloc[_val,:-4].max() * 20)
    sorted_minus.append(question_minimum)
    sorted_plus.append(question_max)



#   Create Workbook. Remove default sheet
wb = Workbook()
s1 = wb["Sheet"]
wb.remove(s1)
    

#   Create the stacked bar chart by clinician
for i in range(0, len(clean_pivot.columns)-4):
    #   Store clinician's ID into temp_name
    #   Store the Average Score (out of 100) into temp_scores
    #   Store the Global Score (out of 100) into temp_glob_scores
    #   Store the Min/Max/std into respective variables
    temp_name = clean_pivot.iloc[:,i].name[1]
    clinician_name_filtered = clean_df[clean_df["Clinician_AD_ID"] == temp_name]
    num_of_responses = len(clinician_name_filtered["Response_ID"].unique())
    
    temp_scores = clean_pivot.iloc[:,i]*20
    temp_scores = round(temp_scores,2)
    

    temp_glob_scores = clean_pivot["Global Mean"]*20
    temp_glob_scores = round(temp_glob_scores,2)

    #   Create DataFrame, transpose and inset a numbered index
    temp_df = pd.DataFrame([temp_scores, temp_glob_scores])
    temp_df = temp_df.T


    temp_df.insert(0, "Questions", question_res.values )
    
    
    #   Remove irrelevant questions, rename columns and create a deep copy
    temp_df.columns = ["Questions","Clinician Average Per Question","Average Across Clinicians"]
    


    temp_quest_list = temp_df.copy()
    
    #   Create clinician sheet, BarChart obect, style and titles
    ws1 = wb.create_sheet(str(temp_name))
    ws1.column_dimensions["B"].width = 100     # pass an integer
    ws1.column_dimensions["C"].width = 20     # pass an integer
    ws1.column_dimensions["D"].width = 20     # pass an integer
    ws1.column_dimensions["E"].width = 20     # pass an integer
    ws1.page_setup.orientation = ws1.ORIENTATION_LANDSCAPE
    ws1.page_setup.fitToHeight = 0
    ws1.page_setup.fitToWidth = 1
    chart1 = StockChart()
    chart1.type = "col"
    chart1.style = 2
    chart1.width = 20
    chart1.title = str(temp_name) + " Set 1"
    chart1.y_axis.title = 'Score'
    chart1.x_axis.title = 'Question ID'
    chart1.y_axis.scaling.min = 0
    chart1.y_axis.scaling.max = 100

    # Convert lists of error values to ErrorBars object
    errorbars = list2errorbars(sorted_plus, sorted_minus, errDir='y')
   
    #   Set the reference of the bar charts, set categories and shape
    #   Reference the avg across clinician data
    chart_data1 = Reference(ws1, min_col = 3, min_row=37, max_row=52 )
    chart_data2 = Reference(ws1, min_col = 4, min_row=37, max_row=52 )

    #   Set Up both series 
    series1 = Series(chart_data1, title="Clinician Average Per Question")
    series2 = Series(chart_data2, title="Average Across Clinicians")

    #   Add error bar to Avg Across Clinician series
    series2.errBars = errorbars

    #   X Axis categories - Question #
    cats = Reference(ws1, min_col=1, min_row=37, max_row=52)

    #   Add series to chart
    chart1.append(series1)
    chart1.append(series2)

    #   Add data to chart
#    chart1.add_data(chart_data)#titles_from_data=True
    chart1.set_categories(cats)


    #   Add the chart to cell "A1", below chart add question list
    ws1.add_chart(chart1, "A1")
    
    #   Set Up the High-level metrics
    ws1["C1"] = "Clinician Average Score"
    ws1["C2"] = round(temp_df["Clinician Average Per Question"].mean(),2)
    ws1["C3"] = "Global Average for this Set"
    ws1["C4"] = round(av_across_set,2)
    ws1["D1"] = "Number of Responses"
    ws1["D2"] = num_of_responses
    ws1["A16"] = "Question list"
    ws1["A16"].font = Font(bold=True)


    temp_quest_list.index = range(1, 17)
    for r in dataframe_to_rows(temp_quest_list, index=True, header=False):
        ws1.append(list(r[:2]))
        
    pts = [NumVal(idx=i) for i in range(len(data) - 1)]
    cache = NumData(pt=pts)
    chart1.series[-1].val.numRef.numCache = cache

    temp_df.index = range(1, 17)
    ws1["A34"] = "Response Data"
    ws1["A34"].font = Font(bold=True)


    #   Below the question list, Add the response data
    for r in dataframe_to_rows(temp_df, index=True, header=True):
        ws1.append(r)


    ws1["A54"] = "Interns Comments"
    ws1["A54"].font = Font(bold=True)

    #   FIRST SHORT ANSWER QUESTION
    clinician_specific_comments = short_answer_sort[short_answer_sort["Clinician_AD_ID"]==temp_name]

    # #   set Question ID and Question to their own variables
    question_ID = clinician_specific_comments.Question_ID.unique()
    
    question_ID_1 = question_ID[1]
    
    question_title = clinician_specific_comments.Question.unique()
    ws1["A55"] = question_title[1]

    clinician_specific_comments_1 = clinician_specific_comments[clinician_specific_comments["Question_ID"]==question_ID_1]
    clinician_specific_comments_1 = clinician_specific_comments_1["Comments"].fillna("N/A")

    for i in range(0, len(clinician_specific_comments_1)):
        ws1["B" + str(56 + i)] = ws1.merge_cells("B"+ str(56 + i) +":D"+ str(56 + i))
        ws1["A" + str(56 + i)] = i + 1
        ws1["B" + str(56 + i)] = clinician_specific_comments_1.iloc[i]
        ws1["B" + str(56 + i)].alignment=Alignment(horizontal='general',
                vertical='top',
                text_rotation=0,
                wrap_text=True,
                shrink_to_fit=True)

        ws1["A" + str(56 + i)].alignment=Alignment(horizontal='general',
                vertical='center',
                text_rotation=0,
                wrap_text=True,
                shrink_to_fit=True)

        if len(clinician_specific_comments_1.iloc[i]) > 300: 
            ws1.row_dimensions[(56 + i)].height = float(75)

        elif len(clinician_specific_comments_1.iloc[i]) > 120: 
            ws1.row_dimensions[(56 + i)].height = float(50)



    next_question = 57 + len(clinician_specific_comments_1)

    ws1["A"+ str(next_question)] = question_title[0]
    question_ID_2 = question_ID[0]
    clinician_specific_comments_2 = clinician_specific_comments[clinician_specific_comments["Question_ID"]==question_ID_2]
    clinician_specific_comments_2 = clinician_specific_comments_2["Comments"].fillna("N/A")

    for i in range(0, len(clinician_specific_comments_2)):
        ws1["B" + str(next_question + 1 + i)] = ws1.merge_cells("B"+ str(next_question + 1 + i) +":D"+ str(next_question + 1 + i))
        ws1["A" + str(next_question + 1 + i)] = i + 1
        ws1["B" + str(next_question + 1 + i)] = clinician_specific_comments_2.iloc[i]
        ws1["B" +str(next_question + 1 + i) ].alignment=Alignment(horizontal='general',
                vertical='top',
                text_rotation=0,
                wrap_text=True,
                shrink_to_fit=True)
        ws1["A" +str(next_question + 1 + i) ].alignment=Alignment(horizontal='general',
                vertical='center',
                text_rotation=0,
                wrap_text=True,
                shrink_to_fit=True)
        if len(clinician_specific_comments_2.iloc[i]) > 300: 
            ws1.row_dimensions[(next_question + 1 + i)].height = float(75)
           
        elif len(clinician_specific_comments_2.iloc[i]) > 120: 
            ws1.row_dimensions[(next_question + 1 + i)].height = float(50)
            

    
# # ##################################################################
# # ##     4                                                         #
# # ##     CONFIRM THAT THE FILE IS SAVING TO THE RIGHT LOCATION     #
# # ##################################################################
wb.save(r"C:\Users\ARidding\Documents\Clevelands\ClevelandResults_Set1_2018-2019--ARTEST.xlsx")
# ##################################################################
